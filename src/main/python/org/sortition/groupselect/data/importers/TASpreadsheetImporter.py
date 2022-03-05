import openpyxl


def importFromSpreadsheet(wb: openpyxl.Workbook, options: dict, numb_preview = None):
    wbsheet = wb.get_sheet_by_name(options['sheet'])

    iHeader = options['startLine']
    iStartRows = options['startLine']+1+options['skipLines']
    iEndRows = wbsheet.max_row-options['ignoreLast']

    if iHeader > iEndRows:
        return [], []

    keys = [cell.value for cell in list(wbsheet.rows)[iHeader-1]]
    vals = []

    if iStartRows < iEndRows:
        if numb_preview is None or iEndRows - iStartRows < numb_preview:
            for i in range(iStartRows, iEndRows+1):
                row = [wbsheet.cell(row=i, column=j+1).value for j in range(len(keys))]
                vals.append(row)
        else:
            for i in range(iStartRows, iStartRows+numb_preview):
                row = [wbsheet.cell(row=i, column=j+1).value for j in range(len(keys))]
                vals.append(row)

    return keys, vals


def __determineDelimiter(file_lines):
    N = 10
    contents = ''.join(file_lines[:N])
    ncs = contents.count(',')
    nss = contents.count(';')
    if (ncs < nss):
        return 'semicolon'
    else:
        return 'comma'
