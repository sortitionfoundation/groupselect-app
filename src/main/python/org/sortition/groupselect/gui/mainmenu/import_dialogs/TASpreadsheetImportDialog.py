import openpyxl

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QVBoxLayout, QRadioButton, QHBoxLayout, QLabel, QPushButton, QWidget, QTableWidget, \
    QFormLayout, QTableWidgetItem, QSpinBox, QErrorMessage, QComboBox

from org.sortition.groupselect.data.importers.TASpreadsheetImporter import importFromSpreadsheet


class TASpreadsheetImportDialog(QtWidgets.QDialog):
    __numbPreview = 10

    __optionDefaults = {
        'startLine': 1,
        'skipLines': 0,
        'ignoreLast': 0,
    }

    __lineEditTypes = {
        'startLine': 'Header at line number',
        'skipLines': 'Number of lines to skip after header',
        'ignoreLast': 'Number of lines to omit at the end',
    }

    __lineEditMins = {
        'startLine': 1,
        'skipLines': 0,
        'ignoreLast': 0,
    }


    def __init__(self, parent, wb: openpyxl.Workbook):
        super(TASpreadsheetImportDialog, self).__init__(parent)

        self.__wb = wb
        self.__optionDefaults['sheet'] = wb.active.title

        self.__setupUi()

        self.__initOptions()

        self.__updateTableContent()

        self.ok = False


    def __setupUi(self):
        # radios delimiter
        radioButtonLayout = QHBoxLayout()
        radioButtonLayout.addWidget(QLabel('Sheet:', self))
        self.sheetBox = QComboBox()
        for sheetID, sheetName in enumerate(self.__wb.sheetnames):
            self.sheetBox.addItem(sheetName, sheetID)
        radioButtonLayout.addWidget(self.sheetBox)

        # line edits
        lineEditsLayout = QFormLayout()
        self.lineEdits = {}
        for editKey, editName in self.__lineEditTypes.items():
            self.lineEdits[editKey] = QSpinBox()
            self.lineEdits[editKey].setMinimum(self.__lineEditMins[editKey])
            lineEditsLayout.addRow(QLabel(f"{editName}:", self), self.lineEdits[editKey])

        # table
        self.__table = QTableWidget()

        # buttons
        self.btn_ok = QPushButton('Ok')
        self.btn_ok.clicked.connect(self.__buttonPress)
        self.btn_cancel = QPushButton('Cancel')
        self.btn_cancel.clicked.connect(self.__buttonPress)
        self.btn_cancel.move(80, 0)

        buttons = QHBoxLayout()
        buttons.addWidget(self.btn_ok)
        buttons.addWidget(self.btn_cancel)
        buttons_widget = QWidget()
        buttons_widget.setLayout(buttons)

        # layout
        layout = QVBoxLayout()
        layout.addWidget(buttons_widget)

        vbox = QVBoxLayout()
        vbox.addSpacing(15)
        vbox.addWidget(QLabel('Please specify how the Spreadsheet should be imported.', self))
        vbox.addLayout(radioButtonLayout)
        vbox.addLayout(lineEditsLayout)
        vbox.addWidget(QLabel('Preview:'))
        vbox.addWidget(self.__table)
        vbox.addWidget(buttons_widget)
        self.setLayout(vbox)


    def __initOptions(self):
        # create options dict from defaults
        self.options = self.__optionDefaults.copy()

        # set values of widgets
        self.sheetBox.setCurrentIndex(self.__wb.sheetnames.index(self.__optionDefaults['sheet']))
        for editKey in self.__lineEditTypes:
            self.lineEdits[editKey].setValue(self.options[editKey])

        # set triggers
        self.sheetBox.currentIndexChanged.connect(self.__widgetUpdated)
        for editKey in self.__lineEditTypes:
            self.lineEdits[editKey].textChanged.connect(self.__widgetUpdated)


    def __updateOptionsFromWidgets(self):
        self.options['sheet'] = self.__wb.sheetnames[self.sheetBox.currentIndex()]

        for editKey in self.__lineEditTypes:
            self.options[editKey] = self.lineEdits[editKey].value()


    def __widgetUpdated(self):
        self.__updateOptionsFromWidgets()
        self.__updateTableContent()


    def __updateTableContent(self):
        keys, vals = importFromSpreadsheet(self.__wb, self.options, numb_preview=self.__numbPreview)
        self.__table.setColumnCount(len(keys))
        self.__table.setHorizontalHeaderLabels(keys)
        self.__table.setRowCount(len(vals))
        for j in range(len(keys)):
            for i in range(len(vals)):
                self.__table.setItem(i, j, QTableWidgetItem(vals[i][j]))


    def __buttonPress(self):
        self.ok = (self.sender() == self.btn_ok)
        if self.ok: self.__updateOptionsFromWidgets()
        self.close()


    @classmethod
    def get_input(cls, parent, fname: str):
        try:
            wb = openpyxl.load_workbook(fname)
        except Exception as e:
            QErrorMessage(parent).showMessage(str(e))
            return

        dialog = cls(parent, wb)
        dialog.exec_()
        if dialog.ok:
            keys, vals = importFromSpreadsheet(wb, dialog.options)
            return dialog.ok, keys, vals, dialog.options
        else:
            return dialog.ok, None, None, {}


    @classmethod
    def get_quick(cls, parent, fname: str, options: dict):
        try:
            wb = openpyxl.load_workbook(fname)
        except Exception as e:
            error_dialog = QErrorMessage(parent)
            error_dialog.showMessage(str(e))
            return

        keys, vals = importFromSpreadsheet(wb, options)
        return True, keys, vals, options
